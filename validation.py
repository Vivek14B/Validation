import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# File path configuration
base_path = "/home/vnrfinance/Downloads/Testing/Daily data validation"
input_path = os.path.join(base_path, "Input file")
data_file = os.path.join(base_path, "VNR_SEEDS_PRIVATE_LIMITEDActual_Expense_Report.xlsx")
exception_file = os.path.join(base_path, "Exception_Report.xlsx")
summary_file = os.path.join(base_path, "Exception_Summary_Report.xlsx")
correction_file = os.path.join(base_path, "Correction_Entries.xlsx")
all_depts_file = os.path.join(base_path, "All_Departments_Output.xlsx")
common_usage_file = os.path.join(base_path, "Common_Usage_Report.xlsx")

# Read input file and extract columns
try:
    df = pd.read_excel(data_file)
except Exception as e:
    raise ValueError(f"Failed to read input file {data_file}: {str(e)}")
df.columns = df.columns.str.strip()
input_columns = df.columns.tolist()

# Load input file to get formatting
try:
    wb_input = load_workbook(data_file)
    ws_input = wb_input.active
except Exception as e:
    raise ValueError(f"Failed to load input file for formatting {data_file}: {str(e)}")
header_style = {
    'font': Font(bold=True),
    'alignment': Alignment(horizontal='left', vertical='center'),
    'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
    'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
}
cell_style = {
    'alignment': Alignment(horizontal='left', vertical='center'),
    'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
}

# Verify column name
if 'Sub Department.Name' not in df.columns:
    raise ValueError(f"Column 'Sub Department.Name' not found in the input file. Available columns: {df.columns.tolist()}")

# Preprocess Sub Department to handle empty-like values
null_like_values = [pd.NA, "N/A", "NaN", "null", "NONE", "NA", "0", "-", "", " ", "\u00A0"]
df['Sub Department.Name'] = df['Sub Department.Name'].replace(null_like_values, "").str.strip()

# Load reference data
ref_files = {
    "FC_Crop": pd.read_excel(os.path.join(input_path, "FC-field crop.xlsx"))['Crop.Name'].dropna().unique(),
    "VC_Crop": pd.read_excel(os.path.join(input_path, "VC-Veg Crop.xlsx"))['Crop.Name'].dropna().unique(),
    "SBFC_Region": pd.read_excel(os.path.join(input_path, "SBFC-Region.xlsx"))['Region.Name'].dropna().unique(),
    "SBVC_Region": pd.read_excel(os.path.join(input_path, "SBVC-Region.xlsx"))['Region.Name'].dropna().unique(),
    "SaleFC_Zone": pd.read_excel(os.path.join(input_path, "SaleFC-Zone.xlsx"))['Zone.Name'].dropna().unique(),
    "SaleVC_Zone": pd.read_excel(os.path.join(input_path, "SaleVC-Zone.xlsx"))['Zone.Name'].dropna().unique(),
    "FC_BU": pd.read_excel(os.path.join(input_path, "FC-BU.xlsx"))['Business Unit.Name'].dropna().unique(),
    "VC_BU": pd.read_excel(os.path.join(input_path, "VC-BU.xlsx"))['Business Unit.Name'].dropna().unique(),
    "Fruit_Crop": pd.read_excel(os.path.join(input_path, "Fruit Crop.xlsx"))['Crop.Name'].dropna().unique(),
    "Common_Crop": pd.read_excel(os.path.join(input_path, "Common crop.xlsx"))['Crop.Name'].dropna().unique(),
    "ProductionFC_Zone": pd.read_excel(os.path.join(input_path, "ProductionFC-Zone.xlsx"))['Zone.Name'].dropna().unique(),
    "ProductionVC_Zone": pd.read_excel(os.path.join(input_path, "ProductionVC-Zone.xlsx"))['Zone.Name'].dropna().unique(),
    "SalesActivity": pd.read_excel(os.path.join(input_path, "SalesActivity.xlsx"))['Activity.Name'].dropna().unique(),
    "MarketingActivity": pd.read_excel(os.path.join(input_path, "MarketingActivity.xlsx"))['Activity.Name'].dropna().unique(),
    "RS_BU": pd.read_excel(os.path.join(input_path, "RS-BU.xlsx"))['Business Unit.Name'].dropna().unique(),
    "SaleRS_Zone": pd.read_excel(os.path.join(input_path, "SaleRS-Zone.xlsx"))['Zone.Name'].dropna().unique(),
    "SBRS_Region": pd.read_excel(os.path.join(input_path, "SBRS-Region.xlsx"))['Region.Name'].dropna().unique(),
    "Root Stock_Crop": pd.read_excel(os.path.join(input_path, "Root Stock Crop.xlsx"))['Crop.Name'].dropna().unique(),
    "Region_Excluded_Accounts": pd.read_excel(os.path.join(input_path, "Region.Name excluded.xlsx"))['Account.Code'].dropna().astype(str).unique(),
    "Zone_Excluded_Accounts": pd.read_excel(os.path.join(input_path, "Zone.Name excluded.xlsx"))['Account.Code'].dropna().astype(str).unique(),   
}

no_crop_check = {
    "Finance & Account", "Human Resource", "Administration",
    "Information Technology", "Legal", "Accounts Receivable & MIS"
}
no_activity_check = no_crop_check.copy()
no_activity_check.add("Production")
no_activity_check.add("Processing")
no_activity_check.add("Parent Seed")

def is_not_blank(value):
    if pd.isna(value) or value is None:
        return False
    val = str(value).strip().replace("\u00A0", "").replace("\u200B", "")
    return val != "" and val.upper() not in ["N/A", "NULL", "NONE", "NA", "0", "-"]

def is_blank(value):
    return not is_not_blank(value)

# Validation logic
def validate_row(dept, row):
    reasons = []
    sub_dept = str(row.get("Sub Department.Name", "") or "").strip().replace("\u00A0", "").replace("\u200B", "")
    func = str(row.get("Function.Name", "") or "").strip()
    vertical = str(row.get("FC-Vertical.Name", "") or "").strip()
    loc = str(row.get("Location.Name", "") or "").strip()
    crop = str(row.get("Crop.Name", "") or "").strip()
    act = str(row.get("Activity.Name", "") or "").strip()
    region = row.get("Region.Name", "")
    zone = row.get("Zone.Name", "")
    bu = row.get("Business Unit.Name", "")
    account_code = str(row.get("Account.Code", "") or "").strip()

    # Generic checks
    if is_blank(loc) or loc.startswith("ZZ"):
        reasons.append("Incorrect Location Name")
    if dept not in no_activity_check and dept not in ["Breeding", "Trialing & PD", "Sales", "Marketing", "Breeding Support"]:
        if is_blank(act) or act.startswith("ZZ"):
            reasons.append("Incorrect Activity Name")
    # Crop and Vertical validation for all departments except those in no_crop_check
    if dept not in no_crop_check:
        if is_blank(vertical):
            reasons.append("FC-Vertical Name cannot be blank")
        if is_blank(crop):
            reasons.append("Crop Name cannot be blank")
        elif crop.startswith("ZZ"):
            reasons.append("Incorrect Crop Name starting with ZZ")
        elif vertical == "FC-field crop" and crop not in ref_files["FC_Crop"]:
            reasons.append("Incorrect Crop Name for FC-field crop Vertical")
        elif vertical == "VC-Veg Crop" and crop not in ref_files["VC_Crop"]:
            reasons.append("Incorrect Crop Name for VC-Veg Crop Vertical")
        elif vertical == "Fruit Crop" and crop not in ref_files["Fruit_Crop"]:
            reasons.append("Incorrect Crop Name for Fruit Crop Vertical")
        elif vertical == "Common" and crop not in ref_files["Common_Crop"]:
            reasons.append("Incorrect Crop Name for Common vertical")
        elif vertical == "Root Stock" and crop not in ref_files["Root Stock_Crop"]:
            reasons.append("Incorrect Crop Name for Root Stock Crop Vertical")
    # Account Code exclusion checks
    if account_code in ref_files["Region_Excluded_Accounts"] and is_not_blank(region):
        reasons.append("Region Name should be blank for this Account Code")
    if account_code in ref_files["Zone_Excluded_Accounts"] and is_not_blank(zone):
        reasons.append("Zone Name should be blank for this Account Code")

    # Department-specific checks
    if dept == "Parent Seed":
        if sub_dept not in ["Breeder Seed Production", "Foundation Seed Production", "Processing FS"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Production":
        if sub_dept not in ["Commercial Seed Production", "Seed Production Research"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        # Zone validation for Commercial Seed Production sub-department
        if sub_dept == "Commercial Seed Production":
            if vertical == "FC-field crop":
                if is_blank(zone):
                    reasons.append("Need to update Zone can not left Blank")
                elif zone not in ref_files["ProductionFC_Zone"]:
                    reasons.append("Incorrect Zone Name for FC-field crop Vertical")
            elif vertical == "VC-Veg Crop":
                if is_blank(zone):
                    reasons.append("Need to update Zone can not left Blank")
                elif zone not in ref_files["ProductionVC_Zone"]:
                    reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
            elif vertical == "Common" and is_blank(zone):
                reasons.append("Need to update Zone Name can not left Blank")

    elif dept == "Processing":
        if sub_dept not in ["Processing", "Warehousing", "Project & Maintenance"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        if loc not in ["Bandamailaram", "Deorjhal", "Boriya"]:
            reasons.append("Need to Update Processing Location")

    elif dept == "Quality Assurance":
        if sub_dept not in ["Field QA", "Lab QC", "Bio Tech Services"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        # Sub-department-specific activity checks
        if sub_dept == "Lab QC" and act not in ["Lab Operations QA", "All Activity"]:
            reasons.append("Incorrect Activity Name for Lab QC")
        if sub_dept == "Field QA" and act not in ["Field Operations QA", "All Activity"]:
            reasons.append("Incorrect Activity Name for Field QA")
        if sub_dept == "Bio Tech Services" and act not in ["Molecular", "All Activity"]:
            reasons.append("Incorrect Activity Name for Bio Tech Services")

    elif dept == "Seed Tech":
        if sub_dept not in ["Aging Test", "Pelleting", "Priming", "Common"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "In Licensing & Procurement":
        if is_not_blank(sub_dept):
            reasons.append("Sub Department should be blank")
        if func != "Supply Chain":
            reasons.append("Incorrect Function Name")
        if vertical in ["", "N/A", "Common"]:
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Breeding":
        if is_not_blank(sub_dept):
            reasons.append("Sub Department should be blank")
        if func != "Research and Development":
            reasons.append("Incorrect Function Name")
        if vertical in ["", "N/A"]:
            reasons.append("Incorrect FC-Vertical Name")
        if dept not in no_activity_check and act not in ["Breeding", "All Activity", "Trialing", "Pre Breeding", "Germplasm Maintainance", "Experimental Seed Production"]:
            reasons.append("Incorrect Activity Name")

    elif dept == "Breeding Support":
        if sub_dept not in ["Pathology", "Biotech - Tissue Culture", "Biotech - Mutation", "Biotech - Markers", "Bioinformatics", "Biochemistry", "Entomology", "Common"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Research and Development":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        # Activity validation: Check for blank or ZZ first, then sub-department-specific checks
        if is_blank(act) or act.startswith("ZZ"):
            reasons.append("Activity Name cannot be blank or start with ZZ")
        else:
            # Sub-department-specific activity checks
            if sub_dept == "Biotech - Markers" and act not in ["Molecular", "Grain Quality", "Seed Treatment", "All Activity"]:
                reasons.append("Incorrect Activity Name for Biotech - Markers")
            elif sub_dept == "Biotech - Tissue Culture" and act not in ["Tissue Culture", "All Activity"]:
                reasons.append("Incorrect Activity Name for Biotech - Tissue Culture")
            elif sub_dept == "Biotech - Mutation" and act not in ["Mutation", "All Activity"]:
                reasons.append("Incorrect Activity Name for Biotech - Mutation")
            elif sub_dept == "Entomology" and act not in ["Entomology", "All Activity"]:
                reasons.append("Incorrect Activity Name for Entomology")
            elif sub_dept == "Pathology" and act not in ["Pathalogy", "All Activity"]:
                reasons.append("Incorrect Activity Name for Pathology")
            elif sub_dept == "Bioinformatics" and act not in ["Bioinformatics", "All Activity"]:
                reasons.append("Incorrect Activity Name for Bioinformatics")
            elif sub_dept == "Biochemistry" and act not in ["Biochemistry", "All Activity"]:
                reasons.append("Incorrect Activity Name for Biochemistry")
            elif sub_dept == "Common" and act not in ["All Activity"]:
                reasons.append("Incorrect Activity Name for Common")

    elif dept == "Trialing & PD":
        if is_not_blank(sub_dept):
            reasons.append("Sub Department should be blank")
        if func != "Research and Development":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        if dept not in no_activity_check and act not in ["CT", "All Activity", "Trialing", "RST"]:
            reasons.append("Incorrect Activity Name")

    elif dept == "Sales":
        valid_subs = ["Sales Brand", "Sales Export", "Sales Institutional & Govt"]
        if sub_dept not in valid_subs:
            reasons.append("Incorrect Sub Department Name")
        if func != "Sales and Marketing":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        # Activity validation for Sales
        if is_blank(act) or act.startswith("ZZ") or act not in ref_files["SalesActivity"]:
            reasons.append("Incorrect Activity Name for Sales")
        # Business Unit, Zone, and Region validation for Sales Brand sub-department
        if sub_dept == "Sales Brand":
            if vertical == "FC-field crop":
                if is_blank(bu):
                    reasons.append("Need to update Business Unit can not left Blank")
                elif bu not in ref_files["FC_BU"]:
                    reasons.append("Incorrect Business Unit Name for FC-field crop Vertical")
                if is_blank(zone):
                    reasons.append("Need to update Zone can not left Blank")
                elif zone not in ref_files["SaleFC_Zone"]:
                    reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                if is_blank(region):
                    reasons.append("Need to update Region Name can not left Blank")
                elif region not in ref_files["SBFC_Region"]:
                    reasons.append("Incorrect Region Name for FC-field crop Vertical")
            elif vertical == "VC-Veg Crop":
                if is_blank(bu):
                    reasons.append("Need to update Business Unit can not left Blank")
                elif bu not in ref_files["VC_BU"]:
                    reasons.append("Incorrect Business Unit Name for VC-Veg Crop Vertical")
                if is_blank(zone):
                    reasons.append("Need to update Zone can not left Blank")
                elif zone not in ref_files["SaleVC_Zone"]:
                    reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                if is_blank(region):
                    reasons.append("Need to update Region Name can not left Blank")
                elif region not in ref_files["SBVC_Region"]:
                    reasons.append("Incorrect Region Name for VC-Veg Crop Vertical")
            elif vertical == "Root Stock":
                if is_blank(bu):
                    reasons.append("Need to update Business Unit can not left Blank")
                elif bu not in ref_files["RS_BU"]:
                    reasons.append("Incorrect Business Unit Name for Root Stock Crop Vertical")
                if is_blank(zone):
                    reasons.append("Need to update Zone can not left Blank")
                elif zone not in ref_files["SaleRS_Zone"]:
                    reasons.append("Incorrect Zone Name for Root Stock Crop Vertical")
                if is_blank(region):
                    reasons.append("Need to update Region Name can not left Blank")
                elif region not in ref_files["SBRS_Region"]:
                    reasons.append("Incorrect Region Name for Root Stock Crop Vertical")

    elif dept == "Marketing":
        valid_subs = ["Business Development", "Digital Marketing", "Product Management"]
        if sub_dept not in valid_subs:
            reasons.append("Incorrect Sub Department Name")
        if func != "Sales and Marketing":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")
        elif vertical == "Root Stock" and any(is_not_blank(x) for x in [region, zone, bu]):
            reasons.append("Region, Zone, BU need to check for Root Stock")
        # Activity validation for Marketing
        if is_blank(act) or act.startswith("ZZ") or act not in ref_files["MarketingActivity"]:
            reasons.append("Incorrect Activity Name for Marketing")

    elif dept == "Finance & Account":
        if sub_dept not in ["Accounts", "Finance", "Analytics, Internal Control & Budget", "Purchase ops", "Secretarial", "Document Management System", "Automation", "Group Company"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Human Resource":
        if sub_dept not in ["Compliances", "HR Ops", "Recruitment", "Team Welfare", "Training", "Common"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Administration":
        if sub_dept not in ["Events", "Maintenance", "Travel Desk","Common"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Information Technology":
        if sub_dept not in ["ERP Support", "Infra & Hardware", "Application Development"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Legal":
        if sub_dept not in ["Compliances", "Litigation","Common"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Accounts Receivable & MIS":
        if sub_dept not in ["Branch and C&F Ops", "Commercial & AR Management", "Common", "Order Processing", "Transport & Logistic"]:
            reasons.append("Incorrect Sub Department Name")
        if func != "Support Functions":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    elif dept == "Management":
        if is_not_blank(sub_dept):
            reasons.append("Sub Department should be blank")
        if func != "Management":
            reasons.append("Incorrect Function Name")
        if is_blank(vertical):
            reasons.append("Incorrect FC-Vertical Name")

    return reasons

def apply_formatting(ws, headers, data_columns):
    # Apply header formatting
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        for attr, style in header_style.items():
            setattr(cell, attr, style)

    # Apply cell formatting
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            for attr, style in cell_style.items():
                setattr(cell, attr, style)
            # Apply number formats based on column
            col_name = headers[col_idx - 1]
            if col_name == "Net amount":
                cell.number_format = '#,##0.00'
            elif col_name in ["Date", "Created date", "Modified date"]:
                cell.number_format = 'DD/MM/YYYY'

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

# Collect data for all departments, exceptions, and common usage
dept_dfs = {}
exception_dfs_dict = {}
common_usage_dfs = {}
for dept in df['Department.Name'].dropna().unique():
    dept_df = df[df['Department.Name'] == dept].copy()
    exceptions = []
    common_usage = []
    for _, row in dept_df.iterrows():
        reasons = validate_row(dept, row)
        if reasons:
            record = row.to_dict()
            record['Exception Reasons'] = "; ".join(reasons)
            exceptions.append(record)
        # Check for "Common" in specified columns
        if (str(row.get("FC-Vertical.Name", "") or "").strip() == "Common" or
            str(row.get("Department.Name", "") or "").strip() == "Common" or
            str(row.get("Sub Department.Name", "") or "").strip() == "Common"):
            common_usage.append(row.to_dict())

    # Prepare dept_df with input columns
    dept_df = dept_df.reindex(columns=input_columns, fill_value='')
    dept_dfs[dept] = dept_df

    # Prepare exceptions with input columns + Exception Reasons
    if exceptions:
        exception_df = pd.DataFrame(exceptions)
        exception_df = exception_df.reindex(columns=input_columns + ['Exception Reasons'], fill_value='')
    else:
        exception_df = pd.DataFrame(columns=input_columns + ['Exception Reasons'])
    exception_dfs_dict[dept] = exception_df

    # Prepare common usage with input columns
    if common_usage:
        common_usage_df = pd.DataFrame(common_usage)
        common_usage_df = common_usage_df.reindex(columns=input_columns, fill_value='')
    else:
        common_usage_df = pd.DataFrame(columns=input_columns)
    common_usage_dfs[dept] = common_usage_df

# Write All_Departments_Output.xlsx
with pd.ExcelWriter(all_depts_file, engine='openpyxl') as writer:
    for dept, dept_df in dept_dfs.items():
        dept_df.to_excel(writer, sheet_name=str(dept)[:31], index=False)
        apply_formatting(writer.book[str(dept)[:31]], input_columns, input_columns)

# Write Exception_Report.xlsx
with pd.ExcelWriter(exception_file, engine='openpyxl') as writer:
    for dept, exception_df in exception_dfs_dict.items():
        exception_df.to_excel(writer, sheet_name=str(dept)[:31], index=False)
        apply_formatting(writer.book[str(dept)[:31]], input_columns + ['Exception Reasons'], input_columns + ['Exception Reasons'])

# Write Common_Usage_Report.xlsx
with pd.ExcelWriter(common_usage_file, engine='openpyxl') as writer:
    for dept, common_usage_df in common_usage_dfs.items():
        common_usage_df.to_excel(writer, sheet_name=str(dept)[:31], index=False)
        apply_formatting(writer.book[str(dept)[:31]], input_columns, input_columns)

# Generate summary report
exception_summary_writer = pd.ExcelWriter(summary_file, engine='openpyxl')

# Read all sheets from Exception_Report.xlsx
try:
    exception_dfs = pd.read_excel(exception_file, sheet_name=None)
except Exception as e:
    raise ValueError(f"Failed to read Exception_Report.xlsx: {str(e)}")

# Filter out empty or all-NA DataFrames
valid_dfs = [df for df in exception_dfs.values() if not df.empty and df.dropna(how='all').shape[0] > 0]
if valid_dfs:
    all_exceptions = pd.concat(valid_dfs, ignore_index=True)
else:
    all_exceptions = pd.DataFrame()

# Handle cases where Exception_Report is empty
if all_exceptions.empty:
    empty_df = pd.DataFrame({'Summary': ['No exceptions found'], 'Count': [0]})
    empty_df.to_excel(exception_summary_writer, sheet_name='Summary', index=False)
    ws = exception_summary_writer.book['Summary']
    apply_formatting(ws, ['Summary', 'Count'], ['Summary', 'Count'])
    exception_summary_writer.close()
else:
    # Split Exception Reasons into individual errors
    all_exceptions['Exception Reasons List'] = all_exceptions['Exception Reasons'].str.split("; ")
    exploded_exceptions = all_exceptions.explode('Exception Reasons List')

    # 1. User-wise Error Summary
    user_summary = pd.pivot_table(
        all_exceptions,
        index=['Created user', 'Modified user'],
        columns='Department.Name',
        values='Exception Reasons',
        aggfunc='count',
        fill_value=0,
        margins=True,
        margins_name='Total'
    )
    user_summary.to_excel(exception_summary_writer, sheet_name='User-wise Summary')
    ws = exception_summary_writer.book['User-wise Summary']
    headers = user_summary.reset_index().columns.tolist()
    apply_formatting(ws, headers, headers)

    # 2. Error Type Summary
    error_type_summary = pd.pivot_table(
        exploded_exceptions,
        index='Exception Reasons List',
        columns='Department.Name',
        values='Created user',
        aggfunc='count',
        fill_value=0,
        margins=True,
        margins_name='Total'
    )
    error_type_summary.to_excel(exception_summary_writer, sheet_name='Error Type Summary')
    ws = exception_summary_writer.book['Error Type Summary']
    headers = error_type_summary.reset_index().columns.tolist()
    apply_formatting(ws, headers, headers)

    # 3. Detailed Error Breakdown
    detailed_summary = pd.pivot_table(
        exploded_exceptions,
        index=['Department.Name', 'Sub Department.Name', 'Created user', 'Exception Reasons List'],
        values=['Net amount'],
        aggfunc={'Net amount': ['count', 'sum']},
        fill_value=0
    )
    detailed_summary.columns = ['Count', 'Total Net Amount']
    detailed_summary = detailed_summary.reset_index()
    detailed_summary.to_excel(exception_summary_writer, sheet_name='Detailed Breakdown', index=False)
    ws = exception_summary_writer.book['Detailed Breakdown']
    headers = detailed_summary.columns.tolist()
    apply_formatting(ws, headers, headers)

    exception_summary_writer.close()

# Generate Correction Entries report
correction_writer = pd.ExcelWriter(correction_file, engine='openpyxl')

# Read all sheets from Exception_Report.xlsx and add Department.Name
corrected_dfs = []
for sheet_name, df in exception_dfs.items():
    if not df.empty and df.dropna(how='all').shape[0] > 0:
        df_copy = df.copy()
        # Ensure Department.Name is included, using sheet name if necessary
        if 'Department.Name' not in df_copy.columns:
            df_copy['Department.Name'] = sheet_name
        else:
            df_copy['Department.Name'] = df_copy['Department.Name'].fillna(sheet_name)
        # Reindex to match input columns + Exception Reasons
        df_copy = df_copy.reindex(columns=input_columns + ['Exception Reasons'], fill_value='')
        corrected_dfs.append(df_copy)

# Concatenate valid DataFrames
if corrected_dfs:
    correction_entries = pd.concat(corrected_dfs, ignore_index=True)
    # Reorder columns to have Department.Name first
    cols = ['Department.Name'] + [col for col in input_columns if col != 'Department.Name'] + ['Exception Reasons']
    correction_entries = correction_entries[cols]
else:
    correction_entries = pd.DataFrame(columns=['Department.Name'] + [col for col in input_columns if col != 'Department.Name'] + ['Exception Reasons'])

# Write to Correction Entries report
correction_entries.to_excel(correction_writer, sheet_name='Correction Entries', index=False)
ws = correction_writer.book['Correction Entries']
apply_formatting(ws, correction_entries.columns.tolist(), correction_entries.columns.tolist())
correction_writer.close()
